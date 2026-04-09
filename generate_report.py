# ================================================
# Automated Marketing Report Generator
# Built by: Margos
# Purpose: Automatically analyze Google Ads campaign
#          data and generate a formatted Excel report
# ================================================

import pandas as pd
import logging
from pathlib import Path
from datetime import datetime
from typing import Optional, Tuple
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# ================================================
# CONFIGURATION
# ================================================
CONFIG = {
    "input_file": "campaign_data.csv",
    "output_file": "campaign_report.xlsx",
    "conversion_value": 800,  # INR per conversion
    "roas_good_threshold": 150,
    "roas_poor_threshold": 80,
}

# ================================================
# SETUP LOGGING
# ================================================
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
)
logger = logging.getLogger(__name__)


# ================================================
# UTILITY FUNCTIONS
# ================================================
def load_campaign_data(file_path: str) -> pd.DataFrame:
    """
    Load campaign data from CSV file with error handling.
    
    Args:
        file_path: Path to the CSV file
        
    Returns:
        DataFrame with campaign data
        
    Raises:
        FileNotFoundError: If the CSV file doesn't exist
        pd.errors.ParserError: If the CSV file is malformed
    """
    try:
        if not Path(file_path).exists():
            raise FileNotFoundError(f"Input file '{file_path}' not found.")
        
        df = pd.read_csv(file_path)
        logger.info(f"Successfully loaded {len(df)} campaigns from {file_path}")
        return df
    
    except FileNotFoundError as e:
        logger.error(str(e))
        raise
    except pd.errors.ParserError as e:
        logger.error(f"Error parsing CSV file: {e}")
        raise


def validate_data(df: pd.DataFrame) -> None:
    """
    Validate that required columns exist and contain valid data.
    
    Args:
        df: DataFrame to validate
        
    Raises:
        ValueError: If required columns are missing or contain invalid data
    """
    required_columns = ["Campaign", "Impressions", "Clicks", "Spend", "Conversions"]
    missing_columns = set(required_columns) - set(df.columns)
    
    if missing_columns:
        raise ValueError(f"Missing required columns: {missing_columns}")
    
    # Check for negative values
    numeric_columns = ["Impressions", "Clicks", "Spend", "Conversions"]
    for col in numeric_columns:
        if (df[col] < 0).any():
            logger.warning(f"Column '{col}' contains negative values")
    
    # Check for division by zero scenarios
    if (df["Impressions"] == 0).any():
        logger.warning("Some campaigns have zero impressions")
    if (df["Clicks"] == 0).any():
        logger.warning("Some campaigns have zero clicks (CPC will be undefined)")
    if (df["Spend"] == 0).any():
        logger.warning("Some campaigns have zero spend (CPC and ROAS will be undefined)")
    
    logger.info("Data validation completed")


def calculate_kpis(df: pd.DataFrame, conversion_value: float) -> pd.DataFrame:
    """
    Calculate Key Performance Indicators (KPIs) for campaign data.
    
    Args:
        df: DataFrame with campaign data
        conversion_value: Value (in INR) per conversion
        
    Returns:
        DataFrame with added KPI columns
    """
    # Click Through Rate (CTR) - Percentage of impressions that resulted in clicks
    df["CTR (%)"] = (df["Impressions"] > 0).apply(
        lambda x: round((df["Clicks"] / df["Impressions"]) * 100, 2) if x else 0
    )
    df["CTR (%)"] = df.apply(
        lambda row: round((row["Clicks"] / row["Impressions"]) * 100, 2) 
        if row["Impressions"] > 0 else 0,
        axis=1
    )
    
    # Cost Per Click (CPC) - Average cost per click in INR
    df["CPC (INR)"] = df.apply(
        lambda row: round(row["Spend"] / row["Clicks"], 2) 
        if row["Clicks"] > 0 else 0,
        axis=1
    )
    
    # Return on Ad Spend (ROAS) - Revenue generated per rupee spent
    df["ROAS"] = df.apply(
        lambda row: round((row["Conversions"] * conversion_value) / row["Spend"], 2)
        if row["Spend"] > 0 else 0,
        axis=1
    )
    
    logger.info("KPIs calculated successfully")
    return df


def get_row_color(roas_value: float, good_threshold: float, poor_threshold: float) -> PatternFill:
    """
    Determine row color based on ROAS performance.
    
    Args:
        roas_value: The ROAS value to evaluate
        good_threshold: ROAS threshold for good performance
        poor_threshold: ROAS threshold for poor performance
        
    Returns:
        PatternFill object with appropriate color
    """
    if roas_value >= good_threshold:
        return PatternFill("solid", fgColor="C6EFCE")  # Light green
    elif roas_value < poor_threshold:
        return PatternFill("solid", fgColor="FFC7CE")  # Light red
    else:
        return PatternFill(fill_type=None)  # No color


def create_excel_report(
    df: pd.DataFrame,
    output_file: str,
    config: dict
) -> None:
    """
    Create and format the Excel report.
    
    Args:
        df: DataFrame with campaign data and KPIs
        output_file: Path to save the Excel file
        config: Configuration dictionary with styling parameters
    """
    # ================================================
    # Create a new Excel Workbook
    # ================================================
    wb = Workbook()
    ws = wb.active
    ws.title = "Campaign Report"
    
    # Define border style
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # ================================================
    # Write the report title and metadata
    # ================================================
    ws["A1"] = "Google Ads Campaign Performance Report"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="1F4E79")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("A1:H1")
    ws.row_dimensions[1].height = 25
    
    ws["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws["A2"].font = Font(italic=True, size=10)
    ws.merge_cells("A2:H2")
    
    ws["A3"] = "Automatically generated using Python"
    ws["A3"].font = Font(italic=True, size=9)
    ws.merge_cells("A3:H3")
    
    # ================================================
    # Write column headers with styling
    # ================================================
    headers = list(df.columns)
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col_num, value=header)
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill("solid", fgColor="1F4E79")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
    
    ws.row_dimensions[5].height = 20
    
    # ================================================
    # Write data rows with color coding
    # Green = Strong Return (ROAS >= threshold)
    # Red   = Poor Return (ROAS < threshold)
    # ================================================
    for row_num, row_data in enumerate(df.values, 6):
        roas_value = row_data[-1]
        row_color = get_row_color(
            roas_value,
            config["roas_good_threshold"],
            config["roas_poor_threshold"]
        )
        
        # Write each cell and apply styling
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = row_color
            cell.border = thin_border
            
            # Format numeric columns
            if col_num > 4:  # Format KPI columns
                if isinstance(value, float):
                    cell.number_format = '0.00'
    
    # ================================================
    # Add summary section
    # ================================================
    summary_row = len(df) + 9
    summary_label_fill = PatternFill("solid", fgColor="D9E1F2")
    
    # Summary header
    ws.cell(row=summary_row, column=1, value="SUMMARY STATISTICS").font = Font(
        bold=True,
        size=12,
        color="FFFFFF"
    )
    ws.cell(row=summary_row, column=1).fill = PatternFill("solid", fgColor="1F4E79")
    ws.merge_cells(f"A{summary_row}:B{summary_row}")
    
    # Summary metrics
    summary_metrics = [
        ("Total Spend (INR):", round(df["Spend"].sum(), 2)),
        ("Total Impressions:", int(df["Impressions"].sum())),
        ("Total Clicks:", int(df["Clicks"].sum())),
        ("Total Conversions:", int(df["Conversions"].sum())),
        ("Average CTR (%):", round(df["CTR (%)"].mean(), 2)),
        ("Average ROAS:", round(df["ROAS"].mean(), 2)),
    ]
    
    for idx, (label, value) in enumerate(summary_metrics, 1):
        label_cell = ws.cell(row=summary_row + idx, column=1, value=label)
        label_cell.fill = summary_label_fill
        label_cell.font = Font(bold=True)
        label_cell.border = thin_border
        
        value_cell = ws.cell(row=summary_row + idx, column=2, value=value)
        value_cell.border = thin_border
        if isinstance(value, float):
            value_cell.number_format = '0.00'
    
    # Best campaign
    best_campaign_idx = df["ROAS"].idxmax()
    best_campaign = df.loc[best_campaign_idx, "Campaign"]
    best_roas = df.loc[best_campaign_idx, "ROAS"]
    
    best_label = ws.cell(row=summary_row + 7, column=1, value="Best Campaign (ROAS):")
    best_label.fill = summary_label_fill
    best_label.font = Font(bold=True)
    best_label.border = thin_border
    
    best_value = ws.cell(row=summary_row + 7, column=2, value=f"{best_campaign} ({best_roas})")
    best_value.border = thin_border
    
    # ================================================
    # Auto-adjust column widths
    # ================================================
    column_widths = {
        "Campaign": 25,
        "Impressions": 18,
        "Clicks": 12,
        "Spend": 15,
        "Conversions": 15,
        "CTR (%)": 12,
        "CPC (INR)": 15,
        "ROAS": 12,
    }
    
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        width = column_widths.get(header, 20)
        ws.column_dimensions[col_letter].width = width
    
    ws.column_dimensions["A"].width = 25
    
    # ================================================
    # Set print options for better printing
    # ================================================
    ws.print_options.horizontalCentered = True
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_margins.left = 0.5
    ws.page_margins.right = 0.5
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5
    
    # ================================================
    # Save the workbook
    # ================================================
    try:
        wb.save(output_file)
        logger.info(f"Report saved successfully to {output_file}")
    except Exception as e:
        logger.error(f"Error saving report: {e}")
        raise


def main():
    """Main execution function."""
    try:
        logger.info("Starting campaign report generation...")
        
        # Load data
        df = load_campaign_data(CONFIG["input_file"])
        
        # Validate data
        validate_data(df)
        
        # Calculate KPIs
        df = calculate_kpis(df, CONFIG["conversion_value"])
        
        # Create Excel report
        create_excel_report(df, CONFIG["output_file"], CONFIG)
        
        # Print summary
        best_campaign = df.loc[df["ROAS"].idxmax(), "Campaign"]
        best_roas = df.loc[df["ROAS"].idxmax(), "ROAS"]
        
        logger.info("=" * 50)
        logger.info("Campaign Report Generated Successfully!")
        logger.info("=" * 50)
        logger.info(f"Output file: {CONFIG['output_file']}")
        logger.info(f"Total campaigns analyzed: {len(df)}")
        logger.info(f"Best Campaign: {best_campaign} (ROAS: {best_roas})")
        logger.info("=" * 50)
        
    except Exception as e:
        logger.error(f"Fatal error: {e}")
        raise


if __name__ == "__main__":
    main()
